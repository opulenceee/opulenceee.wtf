from flask import Flask, render_template, request, send_file, jsonify, flash, redirect, url_for
import os
import tempfile
import pandas as pd
import openpyxl
from datetime import datetime, timedelta, timezone
import calendar
import unicodedata
import re
from werkzeug.utils import secure_filename
import io
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except ImportError:
    from backports.zoneinfo import ZoneInfo  # Fallback for older Python

app = Flask(__name__, 
           template_folder='.', 
           static_folder='..', 
           static_url_path='')
app.config['SECRET_KEY'] = 'mishmarot-2024-super-secret-key-opulenceee-wtf-random-string-12345'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Detect environment
IS_PRODUCTION = os.path.exists('/opt/opulenceee.wtf')

ALLOWED_EXTENSIONS = {'xls', 'xlsx', 'xlsm'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def clean_string(s):
    """Clean strings by removing diacritical marks and non-printable characters"""
    return unicodedata.normalize('NFKD', str(s)).strip()

def extract_day_number(value):
    """Extract the numeric day value from a string"""
    match = re.search(r'\d+', str(value))
    if match:
        return int(match.group())
    return None

def unmerge_and_fill_column_a(wb):
    """Unmerge cells in column A and fill with appropriate day numbers"""
    sheet = wb.active
    
    # Get all merged cell ranges
    merged_cells = list(sheet.merged_cells.ranges)
    
    # Process column A merged cells
    for merged_range in merged_cells:
        if 'A' in str(merged_range):
            # Get the value from the top-left cell of the merged range
            top_left_cell = sheet.cell(row=merged_range.min_row, column=1)
            value = top_left_cell.value
            
            # Unmerge the cells
            sheet.unmerge_cells(str(merged_range))
            
            # Fill all cells in the range with the value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                sheet.cell(row=row, column=1, value=value)
    
    return wb

def get_shift_info(col_idx, actual_date, standby_type=None):
    """Get shift information based on column index"""
    if col_idx == 1:  # Morning shift (Column B in Excel = index 1)
        return {
            'type': 'morning',
            'hebrew': 'משמרת בוקר',
            'start': datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=7),
            'end': datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=15)
        }
    elif col_idx == 2:  # Evening shift (Column C in Excel = index 2)
        return {
            'type': 'evening',
            'hebrew': 'משמרת ערב',
            'start': datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=15),
            'end': datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=23)
        }
    elif col_idx == 3:  # Night shift (Column D in Excel = index 3)
        return {
            'type': 'night',
            'hebrew': 'משמרת לילה',
            'start': datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=23),
            'end': datetime.combine(actual_date + timedelta(days=1), datetime.min.time()) + timedelta(hours=7)
        }
    elif col_idx == 4:  # Miscellaneous shift (Column E in Excel = index 4)
        return {
            'type': 'misc',
            'hebrew': 'שונות',
            'start': datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=9),
            'end': datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=17)
        }
    elif col_idx == 5:  # Standby shift (Column F in Excel = index 5)
        # Set hours based on standby type
        if standby_type == 'בוקר':
            start_hour, end_hour = 7, 15
        elif standby_type == 'ערב':
            start_hour, end_hour = 15, 23
        elif standby_type == 'לילה':
            start_hour, end_hour = 23, 7
        else:
            start_hour, end_hour = 7, 7  # All day standby
            
        hebrew_text = f'כוננות - {standby_type}' if standby_type else 'כוננות כל המשמרות'
        
        if end_hour < start_hour or end_hour == 7:  # Night shift or all-day
            end_time = datetime.combine(actual_date + timedelta(days=1), datetime.min.time()) + timedelta(hours=end_hour)
        else:
            end_time = datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=end_hour)
            
        return {
            'type': 'standby',
            'hebrew': hebrew_text,
            'start': datetime.combine(actual_date, datetime.min.time()) + timedelta(hours=start_hour),
            'end': end_time
        }
    return None

def process_shifts(wb, first_day_date, search_name):
    """Process all shifts in the workbook and return shift details"""
    sheet = wb.active
    shift_details = []
    
    for row_idx in range(2, sheet.max_row + 1):
        # Get day number from column A (index 1)
        day_cell_value = sheet.cell(row=row_idx, column=1).value
        if day_cell_value is None:
            continue
            
        day_number = extract_day_number(day_cell_value)
        if day_number is None:
            continue
            
        actual_date = first_day_date + timedelta(days=day_number - 1)
        
        # Check columns B through F (indices 2-6, but we'll use 1-5 for 0-based indexing)
        for col_idx in range(1, 6):  # Columns B through F (1-5 in 0-based)
            cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value  # +1 because openpyxl is 1-based
            
            if cell_value is not None and clean_string(search_name) in clean_string(str(cell_value)):
                standby_type = None
                
                # For standby shifts (column F, col_idx=5), look for standby type
                if col_idx == 5:  # Standby column
                    # Look backwards in the same column to find standby type
                    current_row = row_idx
                    while current_row >= 2:
                        prev_cell = sheet.cell(row=current_row, column=6).value  # Column F
                        if prev_cell is not None:
                            prev_cell_str = str(prev_cell).strip()
                            if 'בוקר' in prev_cell_str:
                                standby_type = 'בוקר'
                                break
                            elif 'ערב' in prev_cell_str:
                                standby_type = 'ערב'
                                break
                            elif 'לילה' in prev_cell_str:
                                standby_type = 'לילה'
                                break
                        current_row -= 1
                
                shift_info = get_shift_info(col_idx, actual_date, standby_type)
                
                if shift_info:
                    shift_details.append({
                        'day': actual_date.strftime('%d/%m/%Y'),
                        'shift_type': shift_info['type'],
                        'shift_hebrew': shift_info['hebrew'],
                        'start_time': shift_info['start'].strftime('%Y-%m-%d %H:%M:%S'),
                        'end_time': shift_info['end'].strftime('%Y-%m-%d %H:%M:%S')
                    })
    
    return shift_details

def create_ics_content(shift_details):
    """Create an .ics file content from shift details"""
    ics_content = """BEGIN:VCALENDAR
PRODID:-//OPULENCEEE//Shift Scheduler 1.0//EN
X-WR-CALNAME:Work Shifts
X-WR-TIMEZONE:Asia/Jerusalem
VERSION:2.0
CALSCALE:GREGORIAN
METHOD:PUBLISH
"""

    for shift in shift_details:
        israel_tz = ZoneInfo("Asia/Jerusalem")
        start_time = datetime.strptime(shift['start_time'], "%Y-%m-%d %H:%M:%S").replace(tzinfo=israel_tz)
        end_time = datetime.strptime(shift['end_time'], "%Y-%m-%d %H:%M:%S").replace(tzinfo=israel_tz)
        
        start_time_utc = start_time.astimezone(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        end_time_utc = end_time.astimezone(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        dtstamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

        ics_content += f"""BEGIN:VEVENT
DTSTART:{start_time_utc}
DTEND:{end_time_utc}
DTSTAMP:{dtstamp}
CREATED:{dtstamp}
LAST-MODIFIED:{dtstamp}
DESCRIPTION:{shift['shift_hebrew']}\\nהחל מ: {shift['start_time']}\\nעד: {shift['end_time']}
SEQUENCE:0
STATUS:CONFIRMED
SUMMARY:{shift['shift_hebrew']}
TRANSP:OPAQUE
BEGIN:VALARM
TRIGGER:-PT24H
DESCRIPTION:Reminder for {shift['shift_hebrew']}
ACTION:DISPLAY
END:VALARM
BEGIN:VALARM
TRIGGER:-PT1H
DESCRIPTION:Reminder for {shift['shift_hebrew']}
ACTION:DISPLAY
END:VALARM
END:VEVENT
"""

    ics_content += "END:VCALENDAR"
    return ics_content.strip()

@app.route('/mishmarot.css')
def serve_mishmarot_css():
    """Serve the mishmarot-specific CSS file from the current directory"""
    return send_file('mishmarot.css', mimetype='text/css')

@app.route('/script.js')
def serve_mishmarot_js():
    """Serve the mishmarot-specific JavaScript file from the current directory"""
    return send_file('script.js', mimetype='application/javascript')

@app.route('/')
def index():
    current_year = datetime.now().year
    current_month = datetime.now().month
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(current_year - 1, current_year + 3))
    
    return render_template('mishmarot.html', 
                         months=months, 
                         years=years,
                         current_month=current_month,
                         current_year=current_year)

@app.route('/convert', methods=['POST'])
def convert_schedule():
    try:
        # Validate form data
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Please upload .xls, .xlsx, or .xlsm files'}), 400
        
        name = request.form.get('name', '').strip()
        month = int(request.form.get('month'))
        year = int(request.form.get('year'))
        
        if not name:
            return jsonify({'error': 'Please enter your name'}), 400
        
        # Process the file
        first_day_date = datetime(year, month, 1)
        
        # Read the file - support both Excel and HTML files
        try:
            # First, read a small portion to detect file type
            file.seek(0)
            file_start = file.read(100).decode('utf-8', errors='ignore')
            file.seek(0)  # Reset file pointer
            
            # Check if it's HTML content (regardless of extension)
            if '<html' in file_start.lower() or '<!doctype html' in file_start.lower():
                # Handle HTML files like the GUI version
                df = pd.read_html(file, flavor='html5lib')[0]
                # Convert DataFrame to Excel format for processing
                temp_xlsx = io.BytesIO()
                df.to_excel(temp_xlsx, index=False, engine='openpyxl')
                temp_xlsx.seek(0)
                wb = openpyxl.load_workbook(temp_xlsx)
            elif file.filename.lower().endswith('.xls'):
                # Convert XLS to XLSX in memory
                df = pd.read_excel(file, engine='xlrd')
                temp_xlsx = io.BytesIO()
                df.to_excel(temp_xlsx, index=False, engine='openpyxl')
                temp_xlsx.seek(0)
                wb = openpyxl.load_workbook(temp_xlsx)
            else:
                # Handle XLSX and XLSM files
                wb = openpyxl.load_workbook(file)
        except Exception as e:
            return jsonify({'error': f'Error reading file: {str(e)}. Please ensure the file is a valid Excel or HTML file.'}), 400
        
        # Unmerge and fill column A
        wb = unmerge_and_fill_column_a(wb)
        
        # Process shifts
        shift_details = process_shifts(wb, first_day_date, name)
        
        if not shift_details:
            return jsonify({'error': f'No shifts found for {name} in {calendar.month_name[month]} {year}'}), 404
        
        # Create ICS content
        ics_content = create_ics_content(shift_details)
        
        # Create response
        response_data = {
            'shifts': shift_details,
            'ics_content': ics_content,
            'filename': f'shifts_{name}_{month}_{year}.ics'
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500

@app.route('/download_ics', methods=['POST'])
def download_ics():
    try:
        ics_content = request.json.get('ics_content')
        filename = request.json.get('filename', 'shifts.ics')
        
        if not ics_content:
            return jsonify({'error': 'No ICS content provided'}), 400
        
        # Create a temporary file
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.ics', delete=False, encoding='utf-8')
        temp_file.write(ics_content)
        temp_file.close()
        
        return send_file(temp_file.name, as_attachment=True, download_name=filename, mimetype='text/calendar')
        
    except Exception as e:
        return jsonify({'error': f'Error creating download: {str(e)}'}), 500

if __name__ == '__main__':
    import os
    debug_mode = os.environ.get('FLASK_ENV') == 'development'
    app.run(debug=debug_mode, host='127.0.0.1', port=2012)  
